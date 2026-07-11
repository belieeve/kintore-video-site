#!/usr/bin/env python3
"""Google スプレッドシートから動画データを取得してvideos.jsonを生成
スプレッドシートをxlsx形式でダウンロードし、ハイパーリンクの実URL（クリック先）を読み取る。
YouTube動画タイトルを取得し、内容に基づいてカテゴリを自動分類"""

import json
import os
import re
import sys
import time
import tempfile
import urllib.request

import openpyxl

SHEET_ID = '1QF2Cmuds6ao6Y-JTEIdTJIZBWhvRo3Sa_TJ_Amu7GTs'
OUTPUT = 'assets/videos.json'
EXPORT_URL = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'

MENU_SHEET_RE = re.compile(r'^\d+月メニュー$')

# タイトルからカテゴリを自動判定するルール（優先度順）
AUTO_CATEGORY_RULES = [
    ('ボクササイズ', ['ボクササイズ', 'ボクシング', 'キックボクシング', 'パンチ', 'ボクシングエクササイズ']),
    ('ヨガ', ['ヨガ', 'yoga', 'パワーヨガ', 'ヨガフロー', 'ヴィンヤサ']),
    ('ピラティス', ['ピラティス', 'pilates']),
    ('ダンス', ['ダンス', 'dance', 'エアロビクス', 'エアロビ', 'ズンバ', 'zumba']),
    ('HIIT', ['hiit', 'タバタ', 'tabata', 'インターバル']),
    ('ラジオ体操', ['ラジオ体操']),
    ('有酸素', ['有酸素', 'カーディオ', 'cardio', 'ウォーキング', 'ジョギング', 'マラソン', '脂肪燃焼', 'エアロ']),
    ('ストレッチ', ['ストレッチ', 'stretch', '柔軟', 'ほぐし', 'リラックス', 'クールダウン', 'リラクゼーション']),
    ('筋トレ', ['筋トレ', '筋肉', 'トレーニング', 'ワークアウト', 'workout', '腹筋', '背筋', '腕立て',
               'スクワット', 'プランク', 'デッドリフト', 'ベンチプレス', '体幹', 'ダンベル']),
]

# 部位をタイトルから検出
BUI_KEYWORDS = {
    '足': ['足', '脚', 'レッグ', 'ふくらはぎ', '太もも', 'スクワット', 'ランジ'],
    '腹筋': ['腹筋', 'お腹', '腹', 'abs', 'プランク'],
    '背中': ['背中', '背筋', 'バック', '僧帽筋', '広背筋'],
    '胸': ['胸', 'チェスト', '大胸筋', 'ベンチプレス', '胸筋'],
    '腕': ['腕', '二の腕', '上腕', 'アーム', '腕立て'],
    'お尻': ['お尻', 'ヒップ', '臀筋', 'ヒップアップ'],
    '体幹': ['体幹', 'コア', 'core', 'プランク'],
    '肩': ['肩', 'ショルダー', '三角筋'],
    '全身': ['全身', 'フルボディ', 'full body'],
    '上半身': ['上半身', 'アッパー'],
    '下半身': ['下半身', 'ロウワー', '下半身痩せ'],
}


def extract_video_id(url):
    if not url:
        return None
    m = re.search(
        r'(?:youtu\.be/|youtube\.com/watch\?v=|youtube\.com/embed/|youtube\.com/shorts/)([a-zA-Z0-9_-]{11})',
        url,
    )
    return m.group(1) if m else None


def fetch_video_title(video_id):
    """YouTube oEmbed APIで動画タイトルを取得"""
    url = f'https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={video_id}&format=json'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as res:
            data = json.loads(res.read().decode('utf-8'))
            return data.get('title', '')
    except Exception:
        return ''


def auto_categorize(title):
    """タイトルからカテゴリと部位を自動判定"""
    if not title:
        return [], []

    title_lower = title.lower()
    categories = []
    for cat_name, keywords in AUTO_CATEGORY_RULES:
        for kw in keywords:
            if kw.lower() in title_lower:
                if cat_name not in categories:
                    categories.append(cat_name)
                break

    bui_parts = []
    for bui_name, keywords in BUI_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in title_lower:
                if bui_name not in bui_parts:
                    bui_parts.append(bui_name)
                break

    return categories, bui_parts


def download_xlsx():
    """スプレッドシートをxlsx形式でダウンロード（ハイパーリンクの実URLを保持）"""
    req = urllib.request.Request(EXPORT_URL, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=120) as res:
        data = res.read()
    path = os.path.join(tempfile.gettempdir(), 'kintore_sheet_export.xlsx')
    with open(path, 'wb') as f:
        f.write(data)
    return path


def get_url(cell):
    """セルからYouTube URLを取得。表示テキストが古い場合があるため、
    クリック先（ハイパーリンク）を優先する"""
    if cell.hyperlink and cell.hyperlink.target and 'youtu' in str(cell.hyperlink.target):
        return cell.hyperlink.target
    if isinstance(cell.value, str) and 'youtu' in cell.value:
        return cell.value
    return None


def fmt_date(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y%m%d')
    if isinstance(v, (int, float)):
        return str(int(v))
    return str(v)


def add_video(videos, url, category, sub_category='', date='', duration=None):
    if not url or not isinstance(url, str):
        return
    vid = extract_video_id(url)
    if not vid:
        return
    if vid not in videos:
        videos[vid] = {
            'videoId': vid,
            'url': f'https://youtu.be/{vid}',
            'categories': [],
            'dates': [],
            'duration': duration,
        }
    cat_entry = {'category': category, 'subCategory': sub_category}
    if cat_entry not in videos[vid]['categories']:
        videos[vid]['categories'].append(cat_entry)
    if date and date not in videos[vid]['dates']:
        videos[vid]['dates'].append(str(date))
    if duration and isinstance(duration, (int, float)):
        videos[vid]['duration'] = duration


def find_header_row(ws, max_scan=30):
    """ヘッダー行を探す（A列が「日付」、または行内に「◯◯動画URL」ラベルがある行）"""
    for row in ws.iter_rows(min_row=1, max_row=max_scan, max_col=14):
        a = row[0].value
        if isinstance(a, str) and a.strip() == '日付':
            return row[0].row
        for cell in row:
            if isinstance(cell.value, str) and '動画URL' in cell.value:
                return cell.row
    return None


def parse_menu_sheet(videos, ws):
    """◯月メニュー: 日付/部位 + D:ラジオ体操 F:筋トレ H:有酸素① J:有酸素② L,N:ストレッチ"""
    header = find_header_row(ws)
    start = header + 1 if header else 1
    for row in ws.iter_rows(min_row=start):
        vals = [c.value for c in row]
        if not vals:
            continue
        date_str = fmt_date(vals[0])
        bui = str(vals[1]) if len(vals) > 1 and vals[1] else ''

        # 列: D(3)ラジオ体操, E(4)時間, F(5)筋トレ, G(6)時間, H(7)有酸素①, I(8)時間,
        #     J(9)有酸素②, K(10)時間, L(11)ストレッチ, M(12)時間, N(13)ストレッチ, O(14)時間
        col_map = [
            (3, 4, 'ラジオ体操', ''),
            (5, 6, '筋トレ', bui),
            (7, 8, '有酸素', ''),
            (9, 10, '有酸素', ''),
            (11, 12, 'ストレッチ', ''),
            (13, 14, 'ストレッチ', ''),
        ]
        for col_idx, dur_idx, cat, sub in col_map:
            if col_idx < len(row):
                url = get_url(row[col_idx])
                dur = vals[dur_idx] if dur_idx and dur_idx < len(vals) else None
                if url:
                    add_video(videos, url, cat, sub, date_str, dur)


def parse_202312_sheet(videos, ws):
    """202312メニュー: F(5)ラジオ体操(E時間), G(6)メイン動画(Hジャンル, I時間),
    J(9)有酸素(K時間), L(11),M(12)ストレッチ"""
    header = find_header_row(ws)
    start = header + 1 if header else 2
    for row in ws.iter_rows(min_row=start):
        vals = [c.value for c in row]
        if not vals:
            continue
        date_str = fmt_date(vals[0])
        genre = str(vals[7]) if len(vals) > 7 and vals[7] else ''

        col_map = [
            (5, 4, 'ラジオ体操', ''),
            (6, 8, '筋トレ', genre),
            (9, 10, '有酸素', ''),
            (11, None, 'ストレッチ', ''),
            (12, None, 'ストレッチ', ''),
        ]
        for col_idx, dur_idx, cat, sub in col_map:
            if col_idx < len(row):
                url = get_url(row[col_idx])
                dur = vals[dur_idx] if dur_idx is not None and dur_idx < len(vals) else None
                if url:
                    add_video(videos, url, cat, sub, date_str, dur)


def parse_yugata_sheet(videos, ws):
    """夕方メニュー: B(1)ジャンル, D(3)動画URL, E(4)時間（日付なし）"""
    header = find_header_row(ws)
    start = header + 1 if header else 1
    for row in ws.iter_rows(min_row=start):
        vals = [c.value for c in row]
        if len(row) <= 3:
            continue
        genre = str(vals[1]) if len(vals) > 1 and vals[1] else ''
        url = get_url(row[3])
        dur = vals[4] if len(vals) > 4 else None
        if url:
            add_video(videos, url, '夕方メニュー', genre, '', dur)


def parse_radio_sheet(videos, ws):
    """ラジオ体操: A列に動画URLの一覧"""
    for row in ws.iter_rows(min_col=1, max_col=1):
        url = get_url(row[0])
        if url:
            add_video(videos, url, 'ラジオ体操', '')


def main():
    print('スプレッドシートをダウンロード中...')
    xlsx_path = download_xlsx()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    videos = {}
    for ws in wb.worksheets:
        name = ws.title
        if MENU_SHEET_RE.match(name):
            print(f'読み込み中: {name}')
            parse_menu_sheet(videos, ws)
        elif name == '202312メニュー':
            print(f'読み込み中: {name}')
            parse_202312_sheet(videos, ws)
        elif name == '夕方メニュー':
            print(f'読み込み中: {name}')
            parse_yugata_sheet(videos, ws)
        elif name == 'ラジオ体操':
            print(f'読み込み中: {name}')
            parse_radio_sheet(videos, ws)

    video_list = list(videos.values())
    print(f'\n合計動画数: {len(video_list)}')

    # 既存データからタイトルキャッシュを読み込み
    title_cache = {}
    old_count = 0
    if os.path.exists(OUTPUT):
        with open(OUTPUT, 'r', encoding='utf-8') as f:
            old_data = json.loads(f.read())
            old_count = len(old_data)
            for v in old_data:
                if v.get('title'):
                    title_cache[v['videoId']] = v['title']

    # 安全ガード: シート名変更や列構成変更でパース結果が激減した場合は
    # 空・欠損データで上書きせずに失敗させる（本番から動画が全消えするのを防ぐ）
    if not video_list:
        print('エラー: 動画が1本も取得できませんでした。videos.jsonは更新しません。')
        sys.exit(1)
    if old_count and len(video_list) < old_count * 0.8:
        print(f'エラー: 動画数が既存の8割未満に激減 ({old_count}→{len(video_list)})。'
              'シート構成が変わった可能性があるため videos.json は更新しません。')
        sys.exit(1)

    # YouTube動画タイトルを取得（未取得のもののみ）
    new_count = 0
    for i, v in enumerate(video_list):
        vid = v['videoId']
        if vid in title_cache:
            v['title'] = title_cache[vid]
        else:
            print(f'  タイトル取得中 ({i+1}/{len(video_list)}): {vid}')
            v['title'] = fetch_video_title(vid)
            if v['title']:
                new_count += 1
            time.sleep(0.3)  # レート制限対策

    print(f'新規タイトル取得: {new_count}件')

    # タイトルからカテゴリと部位を自動判定
    for v in video_list:
        title = v.get('title', '')
        auto_cats, auto_bui = auto_categorize(title)

        # 自動カテゴリをマージ（スプレッドシートのカテゴリに追加）
        for cat in auto_cats:
            cat_entry = {'category': cat, 'subCategory': ''}
            if cat_entry not in v['categories']:
                v['categories'].append(cat_entry)

        # 自動検出した部位を保存
        if auto_bui:
            v['autoBui'] = auto_bui

    # カテゴリ集計
    cat_count = {}
    for v in video_list:
        for c in v['categories']:
            cat = c['category']
            cat_count[cat] = cat_count.get(cat, 0) + 1
    print(f'カテゴリ別: {json.dumps(cat_count, ensure_ascii=False)}')

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(video_list, f, ensure_ascii=False, indent=2)
    print(f'保存完了: {OUTPUT}')


if __name__ == '__main__':
    main()
