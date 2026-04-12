#!/usr/bin/env python3
"""スプレッドシートからvideos.jsonを更新し、変更があればgit push"""

import openpyxl
import json
import re
import subprocess
import os

XLSX = '/Users/shin/Documents/00_Obsidian/02_開発/103_筋トレ動画サイト/筋トレ記録表.xlsx'
OUTPUT = '/Users/shin/Documents/00_Obsidian/02_開発/103_筋トレ動画サイト/assets/videos.json'
REPO_DIR = '/Users/shin/Documents/00_Obsidian/02_開発/103_筋トレ動画サイト'

def extract_video_id(url):
    if not url:
        return None
    m = re.search(r'(?:youtu\.be/|youtube\.com/watch\?v=|youtube\.com/embed/)([a-zA-Z0-9_-]{11})', url)
    return m.group(1) if m else None

def get_url(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    if cell.value and isinstance(cell.value, str) and 'youtu' in cell.value:
        return cell.value
    return None

def extract_videos():
    wb = openpyxl.load_workbook(XLSX)
    videos = {}

    def add_video(url, category, sub_category="", date="", duration=None):
        if not url:
            return
        vid = extract_video_id(url)
        if not vid:
            return
        if vid not in videos:
            videos[vid] = {
                "videoId": vid,
                "url": f"https://youtu.be/{vid}",
                "categories": [],
                "dates": [],
                "duration": duration
            }
        cat_entry = {"category": category, "subCategory": sub_category}
        if cat_entry not in videos[vid]["categories"]:
            videos[vid]["categories"].append(cat_entry)
        if date and date not in videos[vid]["dates"]:
            videos[vid]["dates"].append(str(date))
        if duration and isinstance(duration, (int, float)):
            videos[vid]["duration"] = duration

    # ◯月メニューシートのみ対象
    menu_sheets = [s for s in wb.sheetnames if re.match(r'^\d+月メニュー$', s)]

    for sheet_name in menu_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            vals = [cell.value for cell in row]
            date_val = vals[0]
            if date_val is None:
                continue
            date_str = str(int(date_val)) if isinstance(date_val, (int, float)) else str(date_val)
            bui = str(vals[1]) if len(vals) > 1 and vals[1] else ""

            col_map = [
                (3, 4, "ラジオ体操", ""),
                (5, 6, "筋トレ", bui),
                (7, 8, "有酸素", ""),
                (9, 10, "有酸素", ""),
                (11, 12, "ストレッチ", ""),
                (13, None, "ストレッチ", ""),
            ]
            for col_idx, dur_idx, cat, sub in col_map:
                if col_idx < len(row):
                    url = get_url(row[col_idx])
                    dur = vals[dur_idx] if dur_idx and dur_idx < len(vals) else None
                    if url:
                        add_video(url, cat, sub, date_str, dur)

    return list(videos.values())

def main():
    if not os.path.exists(XLSX):
        print(f"スプレッドシートが見つかりません: {XLSX}")
        return

    video_list = extract_videos()
    print(f"動画数: {len(video_list)}")

    # JSON書き出し
    new_json = json.dumps(video_list, ensure_ascii=False, indent=2)

    # 既存と比較
    old_json = ""
    if os.path.exists(OUTPUT):
        with open(OUTPUT, 'r', encoding='utf-8') as f:
            old_json = f.read()

    if new_json == old_json:
        print("変更なし。スキップします。")
        return

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(new_json)
    print("videos.json を更新しました。")

    # git commit & push
    os.chdir(REPO_DIR)
    subprocess.run(['git', 'add', 'assets/videos.json'], check=True)
    result = subprocess.run(['git', 'diff', '--cached', '--quiet'])
    if result.returncode == 0:
        print("git上の変更なし。スキップします。")
        return

    subprocess.run(['git', 'commit', '-m', 'chore: スプレッドシートから動画データを自動同期'], check=True)
    subprocess.run(['git', 'push'], check=True)
    print("pushしました。Vercelが自動デプロイします。")

if __name__ == '__main__':
    main()
